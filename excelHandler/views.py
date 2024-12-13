from django.shortcuts import render
from django.http.response import HttpResponse
import datetime
import re
from openpyxl import load_workbook
from .models import *
from glob import glob

def process(request):
    exclude = ['DataEntry', 'constants']
    path = r"C:\Users\User\Jahasiel\HandyScripts\data_extractor\RCI_LFPS_2024.xlsm"
    directory = "C:\\Users\\User\\Jahasiel\\Automations\\RCI Entries\\"
    excel_files = glob(directory + "*.xlsm")

    

    def match_array(find, array):
        for arr in array:
            result = re.search(arr, find, re.IGNORECASE)
            if result:
                return True
            else:
                pass
    
    loaded_wbs = {}
    print("Loading Files...")
    for xl in excel_files:
        loaded_wbs[xl] = load_workbook(xl)
        print(f"Done Loading: {xl}")

    for addr, wb in loaded_wbs.items():
        wb = load_workbook(path, data_only=True)
        for ws in wb.worksheets:

            column_summary_map = {
                'no':1,
                'date':2,
                'check no':3,
                'dv no':4,
                'asa no':5,
                'payee':6,
                'nature of transaction':7,
                'net of tax': 8,
                'gross amount':9
            }
            buffer = []
            if not match_array(ws.title, exclude):
                for cells in ws.iter_cols(min_col=1, max_col=1, min_row=2):
                    for cell in cells:
                        if cell.value:
                            no = ws.cell(row=cell.row, column=column_summary_map['no']).value
                            dte = ws.cell(row=cell.row, column=column_summary_map['date']).value
                            check_no = ws.cell(row=cell.row, column=column_summary_map['check no']).value
                            dv_no = ws.cell(row=cell.row, column=column_summary_map['dv no']).value
                            asa_no = ws.cell(row=cell.row, column=column_summary_map['asa no']).value
                            payee = ws.cell(row=cell.row, column=column_summary_map['payee']).value
                            nature_of_transaction = ws.cell(row=cell.row, column=column_summary_map['nature of transaction']).value
                            net_of_tax = ws.cell(row=cell.row, column=column_summary_map['net of tax']).value
                            gross_amount = ws.cell(row=cell.row, column=column_summary_map['gross amount']).value

                            buffer.append(check_no)

                            def validate_num(param):
                                valid_types = [int, float]
                                if type(param) in valid_types:
                                    return param
                                else:
                                    return 0

                            def validate_date(param):
                                valid_types = [datetime.datetime]
                                
                                if type(param) in valid_types:
                                    return param
                                else:
                                    print(f"Invalid Date Format: {dte}")
                                    dt = input("Correct Format > Y-m-dTH:M:S: ")
                                    return datetime.datetime.fromisoformat(dt)
                                    
                            net_of_tax = validate_num(net_of_tax)
                            gross_amount = validate_num(gross_amount)
                            dte = validate_date(dte)

                            data = {
                                'no':no,
                                'dte':dte,
                                'check_no':check_no,
                                'dv_no':dv_no,
                                'asa_no':asa_no,
                                'payee':payee,
                                'nature_of_transaction':nature_of_transaction,
                                'amountNetOfTax':net_of_tax,
                                'grossAmount':gross_amount
                            }
                            print(no)
                            try:
                                if re.search(r"SDN 501 LFPS RCI", addr):
                                    new_entry = SDN_LFPS_DisbursmentVoucherRecord.objects.create(**data)
                                    new_entry.save()
                                elif re.search(r"SDN 501 COB RCI", addr):
                                    new_entry = SDN_COB_DisbursmentVoucherRecord.objects.create(**data)
                                    new_entry.save()
                                elif re.search(r"SDN 501 CARP RCI", addr):
                                    new_entry = SDN_CARP_DisbursmentVoucherRecord.objects.create(**data)
                                    new_entry.save()
                                elif re.search(r"501 LFPS RCI", addr):
                                    new_entry = ASDI_LFPS_DisbursmentVoucherRecord.objects.create(**data)
                                    new_entry.save()
                                elif re.search(r"501 COB RCI", addr):
                                    new_entry = ASDI_COB_DisbursmentVoucherRecord.objects.create(**data)
                                    new_entry.save()
                                elif re.search(r"501 CARP RCI", addr):
                                    new_entry = ASDI_CARP_DisbursmentVoucherRecord.objects.create(**data)
                                    new_entry.save()
                                else:
                                    pass
                            except Exception as e:
                                print(e)

    return HttpResponse("Done Processing")

# Create your views here.
