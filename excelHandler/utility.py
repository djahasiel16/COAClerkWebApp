import os
from openpyxl import load_workbook
import re
import sqlite3
import datetime

exclude = ['DataEntry', 'constants']
path = input("Path: ")
fund = input("Fund: ")
database_name = input("Database Name: ")
wb = load_workbook(path, data_only=True)


# Adapter function to convert datetime objects to string
def adapt_datetime(dt):
    return dt.strftime("%Y-%m-%d %H:%M:%S")

# Converter function to convert string to datetime objects
def convert_datetime(s):
    return datetime.datetime.strptime(s.decode('utf-8'), "%Y-%m-%d %H:%M:%S")

# Register the adapter and converter
sqlite3.register_adapter(datetime.datetime, adapt_datetime)
sqlite3.register_converter("DATETIME", convert_datetime)


def match_array(find, array):
    for arr in array:
        result = re.search(arr, find, re.IGNORECASE)
        if result:
            return True
        else:
            pass

db = sqlite3.connect(f"{database_name}.sqlite")
for ws in wb.worksheets:

    with open("script.sql", "r") as fr:
        sql_script = fr.read()

    cur = db.cursor()
    cur.executescript(sql_script)

    column_summary_map = {
        'no':1,
        'date':2,
        'check no':3,
        'dv no':4,
        'asa no':5,
        'payee':6,
        'nature of transaction':7,
        'net of tax': 8,
        'gross amount':9
    }
    buffer = []
    if not match_array(ws.title, exclude):
        for cells in ws.iter_cols(min_col=1, max_col=1):
            for cell in cells:
                if cell.value:
                    no = ws.cell(row=cell.row, column=column_summary_map['no']).value
                    dte = ws.cell(row=cell.row, column=column_summary_map['date']).value
                    check_no = ws.cell(row=cell.row, column=column_summary_map['check no']).value
                    dv_no = ws.cell(row=cell.row, column=column_summary_map['dv no']).value
                    asa_no = ws.cell(row=cell.row, column=column_summary_map['asa no']).value
                    payee = ws.cell(row=cell.row, column=column_summary_map['payee']).value
                    nature_of_transaction = ws.cell(row=cell.row, column=column_summary_map['nature of transaction']).value
                    net_of_tax = ws.cell(row=cell.row, column=column_summary_map['net of tax']).value
                    gross_amount = ws.cell(row=cell.row, column=column_summary_map['gross amount']).value

                    buffer.append(check_no)

                    statement = f"INSERT INTO {fund}_summary (no, dte, check_no, dv_no, asa_no, payee, nature_of_transaction, amount_net_of_tax, gross_amount) VALUES (?,?,?,?,?,?,?,?,?)"
                    try:
                        cur.execute(statement, (no, dte, check_no, dv_no, asa_no, payee, nature_of_transaction, net_of_tax, gross_amount))
                    except sqlite3.IntegrityError or check_no in buffer:
                        if check_no:
                            print(f"Last Entry Check No.: {buffer[-1]}")
                            print(f"Current Entry Check No.: {check_no}")
                            while True:
                                try:
                                    check_no = int(input("Provide Correct Check No.: "))
                                    cur.execute(statement, (no, dte, check_no, dv_no, asa_no, payee, nature_of_transaction, net_of_tax, gross_amount))
                                    break
                                except ValueError:
                                    print("Invalid Input!")

               
#####################################################################################################################################

column_map = {
    'check':3,
    'account title':10,
    'dr':11,
    'cr':12
}
entry_data = []
merged_occurance = []
for ws in wb.worksheets:
    for cells in ws.iter_rows(min_col=1, max_col=1):
        for cell in cells:
            for merged_cell in ws.merged_cells:

                if cell.coordinate in merged_cell:
                    if merged_cell not in merged_occurance:
                        merged_occurance.append(merged_cell)
                        entry = []
                        for c in merged_cell.cells:
                            dr = ws.cell(row=c[0], column=column_map['dr']).value
                            cr = ws.cell(row=c[0], column=column_map['cr']).value
                            if dr:
                                entry_data.append({ws.cell(row=cell.row, column=column_map['check']).value:{'account title':ws.cell(row=c[0], column=column_map['account title']).value, 'dr':dr}})
                            elif cr:
                                entry_data.append({ws.cell(row=cell.row, column=column_map['check']).value:{'account title':ws.cell(row=c[0], column=column_map['account title']).value, 'cr':cr}})

                    else:
                        pass
print("Working on Last Phase")
for arr in entry_data:
    for k, v in arr.items():
        statement = f"INSERT INTO {fund}_accounting_entries (check_no, account_title, debit, credit) VALUES (?,?,?,?)"
        try:
            debit = v['dr']
        except KeyError:
            debit = 0
        
        try:
            credit = v['cr']
        except KeyError:
            credit = 0

        try:
            cur.execute(statement,(k, v['account title'], debit, credit))
        except sqlite3.IntegrityError:
            print(arr)
            input("Press Enter to Continue Phase 2...")

    
        os.system("cls")
        print(f"Exectuting {k}")

input("Press Enter to continue... ")
print("Done Last Phase")
db.commit()